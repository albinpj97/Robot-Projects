# +
import openpyxl


from openpyxl import load_workbook


def clear_excel_value(file, option):
    # file = "C:/Users/mafna.janeefar/Documents/Robots/TimesheetWithLogin/Input/extract_template.xlsx"
    # load the work book
    wb_obj = load_workbook(filename = file)
    # select the sheet
    wsheet = wb_obj['Sheet1']
    # sheet.max_row is the maximum number
    # of rows that the sheet have
    # delete_row() method removes rows, first parameter represents row
    # number and sencond parameter represents number of rows
    # to delete from the row number
    if option == '2':
        wsheet.delete_rows(2, wsheet.max_row-1)
    else:
        wsheet.delete_rows(4, wsheet.max_row-1)
  
    # save the file to the path
    wb_obj.save(filename=file)
# -

